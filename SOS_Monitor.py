import socket,os,datetime,queue,threading,time
from openpyxl import Workbook,load_workbook
print("MT: Welcome to SOS Monitoring software")
file=open('settings_irs.txt','r')
data_queue=queue.Queue()
filedic={}
for line in file:
    file_data=line.strip().split('=')
    a=file_data[0]
    b=file_data[1]
    filedic[a]=b
ipaddress_of_system=filedic['ipaddress_of_system']
port_to_listen=int(filedic['port_to_listen'])
filename_of_excel_sheet=filedic['filename_of_excel_sheet']+'.xlsx'
shiftA_start=filedic['shiftA_start_time']
shiftB_start=filedic['shiftB_start_time']
shiftC_start=filedic['shiftC_start_time']
A=list(map(int,shiftA_start.strip().split(":")))
B=list(map(int,shiftB_start.strip().split(":")))
C=list(map(int,shiftC_start.strip().split(":")))
# print(A,B,C)
file.close()
print("MT: Settings file read")

def get_shift(ct):
    startA=datetime.time(A[0],A[1],A[2])
    startB=datetime.time(B[0],B[1],B[2])
    startC=datetime.time(C[0],C[1],C[2])
    if startA<ct<startB:
        return 'A'
    if startB<ct<startC:
        return 'B'
    else:
        return 'C'
# now=datetime.datetime.now()
# print(get_shift(datetime.time(now.hour,now.minute,now.second)))
def excel_data_move():
    while True:
        try:
            mall_flag=False
            data=data_queue.get()
            data=data[2:2+data[1]].decode()
            print(f'EX_TH: Received data: {data}')
            raw_data_list=data.strip().split(",")
            if not os.path.isfile(filename_of_excel_sheet):
                wb=Workbook()
            else:
                wb=load_workbook(filename_of_excel_sheet) 
            if not raw_data_list[0] in wb.sheetnames:
                wb.create_sheet(raw_data_list[0])
            ws=wb[raw_data_list[0]]
            now=datetime.datetime.now()
            date=now.strftime("%d-%m-%Y")
            time_now=now.strftime("%I.%M.%S_%p")
            # print(date,time_to_save)
            dic={}
            dic["DATE"]=date
            dic["SHIFT"]=get_shift(datetime.time(now.hour,now.minute,now.second))
            dic["TIME"]=time_now
            xl_headers=[]
            plc_header=[]
            plc_values=[]
            for i in ws[1]:
                xl_headers.append(i.value)
            for line in raw_data_list[1:]:
                try:
                    temp_list=line.strip().split("-")
                    a=temp_list[0]
                    b=temp_list[1]
                    dic[a]=b
                except:
                    mall_flag=True
                    continue
            for i in dic:
                plc_header.append(i)
                plc_values.append(dic[i])
                
            mc=ws.max_column
            mr=ws.max_row
            for i in plc_header:
                if i not in xl_headers:
                    ws.cell(1,mc+1).value=i
                    xl_headers.append(i)
                    mc=ws.max_column
            for i in xl_headers:
                # print(xlsx_headers.index(i))
                for j in plc_header:
                    if i==j:
                        ws.cell(mr+1,xl_headers.index(i)+1).value=dic[i]

            # worksheet.append(plc_values)
            # print(xl_headers)
            # print(plc_header)
            while True:
                try:
                    wb.save(filename_of_excel_sheet)
                    # print('Saved data : ',dic)
                    if not mall_flag:
                        print(f"EX_TH: Data saved successfully {now}")
                    else:
                        print(f'EX_TH: Received data not in correct format the data not properly recorded {now}')
                    break
                except PermissionError:
                    print("EX_TH: unable to save data in excel sheet, close if it is open...")
                    time.sleep(2)
        except:
            now=datetime.datetime.now()
            print(f"Unhandle error occured! While connection {now}")
                    # continue
excel_data_mo=threading.Thread(target=excel_data_move)
excel_data_mo.start()
print("MT: Excel_data_move thread started")
s=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
s.bind((ipaddress_of_system,port_to_listen))
s.listen()

while True:
    try:
        print("MT: Waiting for data from plc")
        conn,addr=s.accept()
        print(f"MT: Connected to {addr}")
        data=conn.recv(1024)
        # print(data)
        conn.sendall(b'\x09')
        conn.close()
        data_queue.put(data)
    except:
    #     pass
        now=datetime.datetime.now()
        print(f"MT: Unhandled error accquired! While connection {now}")