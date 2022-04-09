from lib2to3.pgen2.pgen import DFAState
import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl import load_workbook
import datetime
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb

def to_excel(df):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': '0.00'}) 
    worksheet.set_column('A:A', None, format1)  
    writer.save()
    processed_data = output.getvalue()
    return processed_data

head=2
st.set_page_config(page_title='IRS_Startup_sheet')
hide_streamlit_style = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
</style>

"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True) 
hide_img_fs = '''
<style>
button[title="View fullscreen"]{
    visibility: hidden;}
</style>
'''

st.markdown(hide_img_fs, unsafe_allow_html=True)

col1, col2 = st.columns((2,1))
# col1,col2 = st.beta_columns([2,1])
logo = Image.open('Logo_ZF_Rane.jpg')
col1.header('IRS Startup Check Sheet')
col2.image(logo, use_column_width=True)
wb =load_workbook('IRS_Startup_sheet.xlsx')
sheet_list = wb.sheetnames
line=st.selectbox('LINE NAME',sheet_list)
report_type=st.selectbox('TYPE',("Shift report","Day report"))
ws=wb[line]
if report_type=="Shift report":
    shift_date=st.date_input("DATE")
    shift_date=shift_date.strftime("%d-%m-%Y")
    shift=st.selectbox("SHIFT",("A","B","C"))
    

if report_type=="Day report":
    day_list=[]
    from_date=st.date_input("FROM DATE")
    max_date=from_date+datetime.timedelta(days=6)
    to_date=st.date_input("TO DATE",value=from_date,min_value=from_date,max_value=max_date)
    delta =  to_date-from_date  

    for i in range(delta.days + 1):
        day = from_date + datetime.timedelta(days=i)
        day_list.append(day.strftime("%d-%m-%Y"))

column=ws.max_column
row=ws.max_row
date_column=1
shift_column=1
required_shift_row=1
a=st.button("Submit")
dic={}
if a:
    show_flag=False
    if report_type=="Shift report": 
        for i in range(1,column+1):
            if ws.cell(row=head,column=i).value=="DATE":
                date_column=i
            if ws.cell(row=head,column=i).value=="SHIFT":
                shift_column=i
        for i in range(1,row+1):
            if ws.cell(row=i,column=date_column).value==shift_date:
                if ws.cell(row=i,column=shift_column).value==shift:
                    required_shift_row=i 
                    show_flag=True
                    # print("r")
        if show_flag:
            header_list=[]
            value_list=[]
            for i in ws[head]:
                if i.value != None:
                    header_list.append(i.value)
                    value_list.append(str(ws.cell(row=required_shift_row,column=i.column).value))
            dic={line:header_list,'CHECK':value_list}
            # st.table(,dic)
            st.table(dic)
            df=pd.DataFrame.from_dict(dic)
            df_xlsx = to_excel(df)
            now=datetime.datetime.now()
            st.download_button(label='📥 Download Current Result',
                                data=df_xlsx ,
                                file_name= 'IRS_SS_{}.xlsx'.format(now.strftime("%d%b%Y_%I-%M-%S%p")))
            # st.
        else:
            st.info("No data found")
    if report_type=="Day report":
        for i in range(1,column+1):
            if ws.cell(row=head,column=i).value=="DATE":
                date_column=i
        count=0
        print(day_list)
        dic={}
        header_list=[]
        header_flag=False
        for i in range(head,row+1):
            if ws.cell(row=i,column=date_column).value in day_list:
                count=count+1
                value_list=[]
                for j in ws[head]:
                    if header_flag==False:
                        header_list.append(j.value)
                    value_list.append(str(ws.cell(row=i,column=j.column).value))
                    show_flag=True
                if header_flag==False:
                    dic[line]=header_list
                    header_flag=True
                dic['Check {}'.format(count)]=value_list
        # print('dic:',dic)
        if show_flag:
            st.table(dic)
            df=pd.DataFrame.from_dict(dic)
            # print(df)
            df_xlsx = to_excel(df)
            now=datetime.datetime.now()
            st.download_button(label='📥 Download Current Result',
                                data=df_xlsx ,
                                file_name= 'IRS_SS_{}.xlsx'.format(now.strftime("%d%b%Y_%I-%M-%S%p")))
        else:
            st.info("No data found")
hide_table_row_index = """
            <style>
            tbody th {display:none}
            .blank {display:none}
            </style>
            """
change_table='''<style>
            table.corner{
                background: red;
            }
            caption{ aifladf}
          thead th { color: white; font-size:14pt;background-color:#FF6700 }
    tr:nth-child(even) {background-color: #f2f2f2;}
    </style>'''
st.markdown(hide_table_row_index, unsafe_allow_html=True)
st.markdown(change_table, unsafe_allow_html=True)
c1,c2=st.columns((3,1))
url = "https://ranegroup.com/rtss"
c2.markdown("Created by [RTSS OSD](%s)" % url)